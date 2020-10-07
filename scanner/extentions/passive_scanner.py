import subprocess
import sys
import re
import netifaces
import openpyxl
import socket
import nmap3
import time
from time import gmtime, strftime
from IPy import IP
from openpyxl.utils import get_column_letter
from extentions.general import get_mac_addr, format_multi_line
from extentions.networking.ethernet import Ethernet
from extentions.networking.ipv4 import IPv4
from extentions.networking.icmp import ICMP
from extentions.networking.tcp import TCP
from extentions.networking.udp import UDP
from extentions.networking.pcap import Pcap
from extentions.networking.http import HTTP

def main():
    select_file = openpyxl.Workbook()
    sheet = select_file.active
    title = ['Mac Address', 'Mac Vendor', 'OS Fingerprint', 'IP Address1', 'IP Address2', 'IP Address3', 'Time1', 'Time2', 'Time3']
    i = 0
    for name in title:
        i = i+1
        title_cell = sheet.cell(row = 1, column = i)
        title_cell.value = name
    adjust_column(sheet)
    select_file.save(filename = 'output.xlsx')
    ver = sys.version
    pattern = re.compile('\A\d{1}.{1}\d{1}.{1}\d{1}.{1}')
    ver = re.findall(pattern,ver)
    print()
    ver = ver[0]
    print('Running Python version ----> %s' %(ver))

   

    def readip():
        """
        create a file named arpscan.txt in the folder where the script is executed.
        In the file enter one ip address range compatable with arp-scan.

        The script will read the file and insert the ip address when
        prompting for an ip address. Simply hit [Enter] to accept the IP.
        You can override the default by typing in an address.
        This allows you to quickly run several different scans with the same
        IP address.
        """

        try:
            IP = []
            f = open('ip.txt', 'r')
            for line in f:
                IP.append(line)
            f.close
        except:  # FileNotFoundError:
            IPAddress = input('Enter an IP Address or network - use /24 style mask: ')
            with open('ip.txt', 'w') as filehandle:
                filehandle.write(IPAddress)
            return IPAddress

        try:
            ipsaved = IP[0]
            ipsaved = ipsaved.strip('\n')
            if not ipsaved:
                IPAddress = input('Enter an IP Address or network - use /24 style mask: ')
            else:
                IPAddress = input('Enter an IP Address or network - use /24 style mask: [%s]: ' % (ipsaved))
                if IPAddress == '':
                    IPAddress = ipsaved
                with open('ip.txt', 'w') as filehandle:
                    filehandle.write(IPAddress)
            if not IPAddress:
                IPAddress = ipsaved

            return IPAddress
        except:
            print('\n[!] An Unknown Error Occured or CTRL+C was pressed')


    def ethinterface():
        """
        Uses the python library netifaces to enumerate a list of the interfaces
        on the system.
        Presents a list of interfaces and prompt the use to select one.
        """
        iflist = netifaces.interfaces()
        print('Interfaces found')

        for index in range(len(iflist)):
            print (index, ':', iflist[index])

        interface = input('enter interface # ')
        interface = int(interface)
        interface = iflist[interface]
    #    interface = input('Enter an interface name if needed: ')
        print('interface selected is:', interface)
        print()
        return interface


    print()
    print('Script usage')
    print('0 Initial arp-scan output')
    print('1 Enter two unused IP addresses')
    print('2 Scan for a lost device')
    print('3 Passive scan')
    print()

    # Make sure 0, 1 or 2 was entered
    scanTest = False
    try:
        while (scanTest != '0' and scanTest != '1'and scanTest != '2' and scanTest != '3'):
            scanTest = input('Input a number to select ')
    except:  # if ctrl+c is pressed exit gracefully
        print('\n[!] An Unknown Error Occured or CTRL+C was pressed')
        raise SystemExit

    scanTest = int(scanTest)
    

    if scanTest == 0:
        # 0 Enter the network and mask to scan Ex. 10.140.100.0/24
        IPAddress = readip()
        #build a list of interfaces and present to the user
        iflist = netifaces.interfaces()
        print('Interfaces found')

        for index in range(len(iflist)):
            print (index, ':', iflist[index])

        interface = input('enter interface # ')
        interface = int(interface)
        interface = iflist[interface]
    #    interface = input('Enter an interface name if needed: ')
        print('interface selected is:',interface)
        print()
        print()
        print('-' * 65)
        print()
        if not interface:
            print('sudo arp-scan --arpspa=127.0.0.1',IPAddress)
            subprocess.run(['sudo', 'arp-scan', '--arpspa=127.0.0.1',IPAddress])
        else:
            I = '-I'
            print('sudo arp-scan',I,interface, '--arpspa=127.0.0.1',IPAddress)
            subprocess.run(['sudo', 'arp-scan',I,interface, '--arpspa=127.0.0.1',IPAddress])
        print()
        print()
        print('-' * 65)
        print()
        if not interface:
            print('sudo arp-scan --arpspa=0.0.0.0',IPAddress)
            subprocess.run(['sudo', 'arp-scan', '--arpspa=0.0.0.0',IPAddress])
        
        else:
            I = '-I'
            print('sudo arp-scan',interface, '--arpspa=0.0.0.0',IPAddress)
            result = subprocess.run(['sudo', 'arp-scan',I,interface, '--arpspa=0.0.0.0',IPAddress],universal_newlines = True ,stdout=subprocess.PIPE).stdout.splitlines()
            find_device(result)

        print()
        print()
        print('-' * 65)
        print()
        if not interface:
            print('sudo arp-scan --arpspa=255.255.255.255', IPAddress)
            subprocess.run(['sudo', 'arp-scan', '--arpspa=255.255.255.255', IPAddress])
        else:
            I = '-I'
            print('sudo arp-scan', I, interface, '--arpspa=255.255.255.255', IPAddress)
            subprocess.run(['sudo', 'arp-scan', I, interface,
                           '--arpspa=255.255.255.255', IPAddress])
        print()
        print()
        print('-' * 65)
        print()
        if not interface:
            print('sudo arp-scan --arpspa=1.0.0.1', IPAddress)
            subprocess.run(['sudo', 'arp-scan', '--arpspa=1.0.0.1', IPAddress])
        else:
            I = '-I'
            print('sudo arp-scan', interface, '--arpspa=1.0.0.1', IPAddress)
            subprocess.run(['sudo', 'arp-scan', I, interface,
                           '--arpspa=1.0.0.1', IPAddress])
        print()
        print()
        print('-' * 65)


    elif scanTest == 1:
        try:
        # 1 Enter the two IP addresses to test
            IPAddress = input('Enter the 1st IP Address ')
            IPAddress1 = input('Enter the 2nd IP Address ')
            interface = ethinterface()
        #interface = input('Enter an interface name if needed ')
        #  print some space
            print()
            print()
            print('-' * 65)
        #  If no interface is entered
            if not interface:
                print('not interface')
                print('sudo arp-scan --arpspa='+IPAddress, IPAddress1)
                print('sudo arp-scan --arpspa='+IPAddress1, IPAddress)
                arp = '--arpspa=' + IPAddress
                arp1 = '--arpspa=' + IPAddress
                subprocess.run(['sudo', 'arp-scan', arp, IPAddress1])
                subprocess.run(['sudo', 'arp-scan', arp1, IPAddress])
    # If an interface is entered
            else:
                I = '-I'
                arp = '--arpspa=' + IPAddress
                arp1 = '--arpspa=' + IPAddress1
                print('sudo arp-scan', I, interface, '--arpspa='+IPAddress, IPAddress1)
                print('sudo arp-scan', I, interface, '--arpspa='+IPAddress1, IPAddress)
                print('-' * 65)
                print()
                subprocess.run(['sudo', 'arp-scan', I, interface, arp, IPAddress1])
                subprocess.run(['sudo', 'arp-scan', I, interface, arp1, IPAddress])
            print()
            print('-' * 65)
            print()
        except:
            print('\n[!] An Unknown Error Occured or CTRL+C was pressed')

    elif scanTest == 2:
#    try:
        
        vlan_ID = input('Enter the vlan ID[default:null]: ')
        dest_MAC_addr = input('Enter the MAC Address: ')
        IPAddress = input('Enter the IP Subnet[default:localnet]: ')
        interface = ethinterface()
#    interface = input('Enter an interface: ')
    #  print some space
        print()
        print(f'IP Subnet {IPAddress}')
        print('-' * 65)
        if interface:
            print(f'To re-run copy/paste this line:')
            print(f'sudo arp-scan -I {interface} -Q {vlan_ID} --destaddr={dest_MAC_addr} {IPAddress}')
            print()
            
            if IPAddress == '':
                IPAddress = '--localnet'
            mac_args = '--destaddr=' + dest_MAC_addr
            command = ['sudo', 'arp-scan', '-I', interface, mac_args, IPAddress]
            if vlan_ID :
                vlan_args = '--vlan=' + vlan_ID
                command.append(vlan_args)
            print(command)
            subprocess.run(command)
            

        #  remove comments for debugging
        #  print(f'MAC Addr {dest_MAC_addr}')
        #  print(f'arg={args}')
            
            

            print()
            print('-' * 65)
            print()
#    except:
#        print('\n[!] An Unknown Error Occured or CTRL+C was pressed')
    elif scanTest == 3:
        passive_scan()

def passive_scan():
    TAB_1 = '\t - '
    TAB_2 = '\t\t - '
    TAB_3 = '\t\t\t - '
    TAB_4 = '\t\t\t\t - '

    DATA_TAB_1 = '\t   '
    DATA_TAB_2 = '\t\t   '
    DATA_TAB_3 = '\t\t\t   '
    DATA_TAB_4 = '\t\t\t\t   '
    
    pcap = Pcap('capture.pcap')
    conn = socket.socket(socket.AF_PACKET, socket.SOCK_RAW, socket.ntohs(3))
    
    exception_ip = ['0.0.0.0', '127.0.0.1', '192.168.255.1', '10.100.32.74']
    exception_ttl = ['1']

    while True:
        try:
            raw_data, addr = conn.recvfrom(65535)
            pcap.write(raw_data)
            eth = Ethernet(raw_data)

            print('\nEthernet Frame:')
            print(TAB_1 + 'Destination: {}, Source: {}, Protocol: {}'.format(eth.dest_mac, eth.src_mac, eth.proto))
            ipv4 = IPv4(eth.data) 
            ip_address = IP(ipv4.src)
            ip_type = ip_address.iptype()
            if ip_type == 'PRIVATE':
                if ipv4.src not in exception_ip:
                    if str(ipv4.ttl) not in exception_ttl:
                        #IPv4
                        if eth.proto == 8:
                            #ipv4 = IPv4(eth.data)
                            print(TAB_1 + 'IPv4 Packet:')
                            print(TAB_2 + 'Version: {}, Header Length: {}, TTL: {},'.format(ipv4.version, ipv4.header_length, ipv4.ttl))
                            print(TAB_2 + 'Protocol: {}, Source: {}, Target: {}'.format(ipv4.proto, ipv4.src, ipv4.target))
                        
                        
                            open_excel(eth.src_mac,ipv4.src,'1')
        except KeyboardInterrupt:
            open_excel(eth.src_mac,ipv4.src,'1')
            #select_file.save('output.xlsx')
            time.sleep(5)
            print('>>>>>>>>>>>>>>>>>>>>>>>.save file.<<<<<<<<<<<<<<<<<<<<<<<<<<<<,,,')
            break
    pcap.close()

def find_device(result):
    result_list = result[2:-4]
    for i in result_list:
            ip = re.search(r'^(.*?)\t', i).group()
            mac = re.search(r'\t(.*?)\t', i).group()
            vendor = re.search(r'([^\t]+)$', i).group()
            open_excel(mac,ip,vendor)
            print(ip,mac,vendor)
            

def open_excel(src_mac,src_ip,mac_vendor):
    select_file = openpyxl.load_workbook('output.xlsx')
    print('Open Excel')
    #select_file = openpyxl.Workbook()
    sheet = select_file.active
    find_mac = 0
    for row in range(1, sheet.max_row + 1):
        mac_cell = sheet.cell(row = row, column = 1)
        cell_value = mac_cell.value
        if cell_value == src_mac:
           find_mac = 1
           check_ip(row,sheet,src_ip)

    if find_mac == 0:
        new_mac = sheet.cell(row = sheet.max_row+1, column = 1)
        print(sheet.max_row+1)
        new_mac_vendor = sheet.cell(row = sheet.max_row, column = 2)
        new_os = sheet.cell(row = sheet.max_row, column = 3)
        new_mac.value = src_mac
#        new_mac_vendor.value = mac_vendor
#        return_os = check_ip(sheet.max_row,sheet,src_ip)
        check_ip(sheet.max_row,sheet,src_ip)
#       new_os.value = return_os
    
    #adjust_column(sheet)
    
    select_file.save('output.xlsx')
#    print('>>>>>>>>>>>>>>>>>>>>>>>Save & Close File<<<<<<<<<<<<<<<<<<<<<<<<<<<<')
def os_scan(ip):
    nmap = nmap3.Nmap()
    os_results = nmap.nmap_os_detection(ip)
    if os_results:
        os_list = (os_results[-1])
        print(os_list['cpe'])
        return(os_list['cpe'])

def check_ip(row,sheet,src_ip):
    #for i in range(4 , sheet.max_column+1):
    for i in range(4 , 7):
        scan_time = time.ctime(time.time())
        ip_cell = sheet.cell(row = row, column = i)
        cell_value = ip_cell.value
        time_cell = sheet.cell(row = row, column = i+3)
        time_cell.value = scan_time
        if cell_value == src_ip:
            break

        elif cell_value is None:
            ip_cell.value = src_ip
            break
#            os = os_scan(src_ip)
#            return(os)
#        elif cell_value == src_ip :
#            os = os_scan(src_ip)
#            return(os) 


def adjust_column(sheet):
    for col in range(1,10):
        column_name = get_column_letter(col)
        sheet.column_dimensions[column_name].auto_size = True
            

main()